import os
import uuid
from functools import wraps
from datetime import datetime, timedelta

from flask import (
    Flask, render_template, request, redirect, url_for,
    flash, jsonify, send_from_directory, abort, Response, session
)
from flask_login import (
    LoginManager, login_user, logout_user, login_required, current_user
)
from dotenv import load_dotenv

from models import db, User, Observation, VALID_CATEGORIES, VALID_SEVERITIES
import groq_service
import drive_service
import departments
import email_service

load_dotenv()

BASE_DIR = os.path.abspath(os.path.dirname(__file__))
DATA_DIR = os.environ.get("DATA_DIR", BASE_DIR)
UPLOAD_DIR = os.path.join(DATA_DIR, "uploads")
os.makedirs(UPLOAD_DIR, exist_ok=True)
os.makedirs(os.path.join(DATA_DIR, "instance"), exist_ok=True)

app = Flask(__name__)
app.config["SECRET_KEY"] = os.environ.get("SECRET_KEY", "dev-secret-change-me")
app.config["SQLALCHEMY_DATABASE_URI"] = "sqlite:///" + os.path.join(DATA_DIR, "instance", "safety.db")
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
app.config["MAX_CONTENT_LENGTH"] = 25 * 1024 * 1024

app.config["SESSION_COOKIE_SECURE"] = True
app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"
app.config["PERMANENT_SESSION_LIFETIME"] = timedelta(hours=8)

db.init_app(app)

login_manager = LoginManager(app)
login_manager.login_view = "login"

_failed_logins = {}
MAX_ATTEMPTS = 5
LOCKOUT_MINUTES = 15


def _is_locked(username):
    entry = _failed_logins.get(username)
    if entry and entry.get("locked_until") and datetime.utcnow() < entry["locked_until"]:
        return entry["locked_until"]
    return None


def _record_failed_login(username):
    entry = _failed_logins.setdefault(username, {"count": 0, "locked_until": None})
    entry["count"] += 1
    if entry["count"] >= MAX_ATTEMPTS:
        entry["locked_until"] = datetime.utcnow() + timedelta(minutes=LOCKOUT_MINUTES)
        entry["count"] = 0


def _clear_failed_login(username):
    _failed_logins.pop(username, None)


@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))


def roles_required(*roles):
    def decorator(fn):
        @wraps(fn)
        def wrapper(*args, **kwargs):
            if not current_user.is_authenticated:
                return redirect(url_for("login"))
            if current_user.role not in roles:
                abort(403)
            return fn(*args, **kwargs)
        return wrapper
    return decorator


@app.route("/", methods=["GET"])
def index():
    if current_user.is_authenticated:
        return redirect(url_for("dashboard_redirect"))
    return redirect(url_for("login"))


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")

        locked_until = _is_locked(username)
        if locked_until:
            minutes_left = max(1, int((locked_until - datetime.utcnow()).total_seconds() // 60) + 1)
            flash(f"Too many failed attempts. Try again in {minutes_left} minute(s).", "error")
            return render_template("login.html")

        user = User.query.filter_by(username=username).first()
        if user and user.check_password(password):
            _clear_failed_login(username)
            session.permanent = True
            login_user(user)
            return redirect(url_for("dashboard_redirect"))
        _record_failed_login(username)
        flash("Invalid username or password.", "error")
    return render_template("login.html")


@app.route("/logout")
@login_required
def logout():
    logout_user()
    return redirect(url_for("login"))


@app.route("/dashboard")
@login_required
def dashboard_redirect():
    if current_user.role == "admin":
        return redirect(url_for("admin_dashboard"))
    if current_user.role == "hse":
        return redirect(url_for("hse_dashboard"))
    return redirect(url_for("user_dashboard"))


@app.route("/user")
@login_required
def user_dashboard():
    my_observations = (
        Observation.query.filter_by(reporter_id=current_user.id)
        .order_by(Observation.created_at.desc())
        .all()
    )
    return render_template("user_dashboard.html", observations=my_observations)


def _notify_department(obs):
    """Look up the manager for this observation's location, save the
    department, and email them. Wrapped so it can never crash the request —
    the observation is already saved by the time this runs."""
    try:
        info = departments.get_location_info(obs.location or "")
        dept = info["department"] if info else None
        mgr = info["manager"] if info else None
        mgr_email = info["manager_email"] if info else None

        if dept:
            obs.department = dept
            db.session.commit()

        if mgr_email:
            print(f"--- ATTEMPTING EMAIL TO: {mgr_email} FOR LOCATION: {obs.location} ---")
            success = email_service.send_observation_email(obs.to_dict(), dept, mgr, mgr_email)
            print(f"--- EMAIL RESULT: {success} ---")
    except Exception as e:
        app.logger.warning(f"Department notification failed: {e}")
        print(f"--- DEPARTMENT NOTIFICATION ERROR: {e} ---")


def _save_audio_and_process(audio_file, fallback_reporter_name=""):
    ext = os.path.splitext(audio_file.filename)[1] or ".webm"
    filename = f"{uuid.uuid4().hex}{ext}"
    local_path = os.path.join(UPLOAD_DIR, filename)
    audio_file.save(local_path)

    result = groq_service.process_audio_report(
        local_path, fallback_reporter_name=fallback_reporter_name
    )

    drive_id, drive_link = None, None
    try:
        drive_id, drive_link = drive_service.upload_audio(local_path, filename)
    except Exception as e:
        app.logger.warning(f"Drive upload failed: {e}")

    return result, filename, drive_id, drive_link


@app.route("/api/webhook/debug", methods=["GET", "POST"])
def webhook_debug():
    info = {
        "method": request.method,
        "headers": dict(request.headers),
        "form_fields": {k: v for k, v in request.form.items()},
        "files_received": [
            {"field_name": k, "filename": f.filename, "content_type": f.content_type}
            for k, f in request.files.items()
        ],
        "raw_body_length": len(request.get_data()) if not request.form and not request.files else None,
    }
    app.logger.info(f"WEBHOOK DEBUG: {info}")
    print(f"--- Incoming webhook-debug request ---\n{info}\n")
    return jsonify(info), 200


@app.route("/api/webhook/voice-observation", methods=["POST"])
def webhook_voice_observation():
    expected_key = os.environ.get("WEBHOOK_API_KEY")
    if expected_key:
        provided_key = (
            request.headers.get("X-Webhook-Key")
            or request.form.get("api_key")
            or request.form.get("secret")
        )
        if provided_key != expected_key:
            return jsonify({"error": "Unauthorized"}), 401

    audio_file = request.files.get("audio") or request.files.get("file")
    if not audio_file or audio_file.filename == "":
        return jsonify({"status": "ok", "note": "Connected. No audio in this request."}), 200

    reporter_name = request.form.get("reporter_name", "").strip()
    reporter_username = request.form.get("username", "").strip()

    reporter = None
    if reporter_username:
        reporter = User.query.filter_by(username=reporter_username).first()

    try:
        result, filename, drive_id, drive_link = _save_audio_and_process(
            audio_file, fallback_reporter_name=reporter_name or (reporter.name if reporter else "")
        )
    except Exception as e:
        import traceback
        app.logger.error("WEBHOOK PROCESSING FAILED:\n" + traceback.format_exc())
        print("WEBHOOK PROCESSING FAILED:\n" + traceback.format_exc())
        return jsonify({"error": f"AI processing failed: {e}"}), 500

    obs = Observation(
        reporter_id=reporter.id if reporter else None,
        reporter_name=result["reporter_name"] or reporter_name or (reporter.name if reporter else "Anonymous"),
        category=result["category"],
        severity=result["severity"],
        location=result["location"],
        urdu_script=result["urdu_script"],
        english_translation=result["english_translation"],
        audio_filename=filename,
        drive_file_id=drive_id,
        drive_link=drive_link,
        status="pending",
    )
    db.session.add(obs)
    db.session.commit()

    _notify_department(obs)

    return jsonify({"success": True, "observation": obs.to_dict()}), 201


@app.route("/hse")
@roles_required("hse", "admin")
def hse_dashboard():
    return render_template("hse_dashboard.html", role="hse")


@app.route("/hse/observations/<int:obs_id>", methods=["PUT"])
@roles_required("hse", "admin")
def hse_update_observation(obs_id):
    obs = Observation.query.get_or_404(obs_id)
    data = request.get_json(force=True)
    for field in ("category", "severity", "location", "status", "reporter_name"):
        if field in data:
            setattr(obs, field, data[field])
    obs.updated_at = datetime.utcnow()
    db.session.commit()
    return jsonify(obs.to_dict())


@app.route("/admin")
@roles_required("admin")
def admin_dashboard():
    users = User.query.order_by(User.created_at.desc()).all()
    return render_template("admin_dashboard.html", role="admin", users=users)


@app.route("/admin/observations/<int:obs_id>", methods=["PUT"])
@roles_required("admin")
def admin_update_observation(obs_id):
    obs = Observation.query.get_or_404(obs_id)
    data = request.get_json(force=True)
    for field in ("category", "severity", "location", "status", "reporter_name",
                  "urdu_script", "english_translation"):
        if field in data:
            setattr(obs, field, data[field])
    obs.updated_at = datetime.utcnow()
    db.session.commit()
    return jsonify(obs.to_dict())


@app.route("/admin/observations/<int:obs_id>", methods=["DELETE"])
@roles_required("admin")
def admin_delete_observation(obs_id):
    obs = Observation.query.get_or_404(obs_id)
    if obs.drive_file_id:
        drive_service.delete_audio(obs.drive_file_id)
    if obs.audio_filename:
        local_path = os.path.join(UPLOAD_DIR, obs.audio_filename)
        if os.path.exists(local_path):
            os.remove(local_path)
    db.session.delete(obs)
    db.session.commit()
    return jsonify({"deleted": True})


@app.route("/admin/observations", methods=["POST"])
@roles_required("admin")
def admin_create_observation():
    data = request.get_json(force=True)
    obs = Observation(
        reporter_id=None,
        reporter_name=data.get("reporter_name", "Unknown"),
        category=data.get("category", "Not specified"),
        severity=data.get("severity", "Not specified"),
        location=data.get("location", "Not specified"),
        urdu_script=data.get("urdu_script", ""),
        english_translation=data.get("english_translation", ""),
        status="pending",
    )
    db.session.add(obs)
    db.session.commit()
    _notify_department(obs)
    return jsonify(obs.to_dict()), 201


@app.route("/admin/users", methods=["POST"])
@roles_required("admin")
def admin_create_user():
    data = request.get_json(force=True)
    if User.query.filter_by(username=data.get("username")).first():
        return jsonify({"error": "Username already exists"}), 400
    user = User(
        name=data.get("name", ""),
        username=data.get("username", ""),
        role=data.get("role", "user"),
    )
    user.set_password(data.get("password", "changeme123"))
    db.session.add(user)
    db.session.commit()
    return jsonify(user.to_dict()), 201


@app.route("/admin/users/<int:user_id>", methods=["DELETE"])
@roles_required("admin")
def admin_delete_user(user_id):
    if user_id == current_user.id:
        return jsonify({"error": "You cannot delete your own account"}), 400
    user = User.query.get_or_404(user_id)
    db.session.delete(user)
    db.session.commit()
    return jsonify({"deleted": True})


@app.route("/api/observations")
@roles_required("hse", "admin")
def api_observations():
    q = Observation.query
    category = request.args.get("category")
    severity = request.args.get("severity")
    reporter = request.args.get("reporter")
    department = request.args.get("department")
    if category:
        q = q.filter(Observation.category == category)
    if severity:
        q = q.filter(Observation.severity == severity)
    if reporter:
        q = q.filter(Observation.reporter_name == reporter)
    if department:
        q = q.filter(Observation.department == department)
    observations = q.order_by(Observation.created_at.desc()).all()
    return jsonify([o.to_dict() for o in observations])


@app.route("/api/observations/<int:obs_id>/audio")
@login_required
def api_get_audio(obs_id):
    obs = Observation.query.get_or_404(obs_id)
    if current_user.role == "user" and obs.reporter_id != current_user.id:
        abort(403)
    if obs.drive_link:
        return redirect(obs.drive_link)
    if obs.audio_filename:
        return send_from_directory(UPLOAD_DIR, obs.audio_filename)
    abort(404)


@app.route("/api/observations/export.csv")
@roles_required("hse", "admin")
def export_csv():
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    observations = Observation.query.order_by(Observation.created_at.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Observations"

    headers = ["TIME", "REPORTER", "CATEGORY", "SEVERITY", "LOCATION", "DEPARTMENT",
               "URDU SCRIPT", "ENGLISH TRANSLATION", "STATUS"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

    for o in observations:
        pkt_time = (o.created_at + timedelta(hours=5)) if o.created_at else None
        ws.append([
            pkt_time.strftime("%d %b, %H:%M") if pkt_time else "",
            o.reporter_name or "",
            o.category or "",
            o.severity or "",
            o.location or "",
            o.department or "",
            o.urdu_script or "",
            o.english_translation or "",
            o.status or "",
        ])

    ws.auto_filter.ref = ws.dimensions

    widths = [16, 18, 16, 12, 16, 14, 40, 45, 12]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return Response(
        buf.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=observation_log.xlsx"},
    )


@app.cli.command("init-db")
def init_db():
    db.create_all()
    if not User.query.filter_by(username="admin").first():
        admin = User(name="Site Admin", username="admin", role="admin")
        admin.set_password("admin123")
        db.session.add(admin)
    if not User.query.filter_by(username="hse").first():
        hse = User(name="HSE Officer", username="hse", role="hse")
        hse.set_password("hse123")
        db.session.add(hse)
    if not User.query.filter_by(username="worker").first():
        worker = User(name="Site Worker", username="worker", role="user")
        worker.set_password("worker123")
        db.session.add(worker)
    db.session.commit()
    print("Database initialised.")
    print("Admin login:  admin / admin123")
    print("HSE login:    hse / hse123")
    print("User login:   worker / worker123")


if __name__ == "__main__":
    with app.app_context():
        os.makedirs(os.path.join(BASE_DIR, "instance"), exist_ok=True)
        db.create_all()
        if not User.query.filter_by(username="admin").first():
            admin = User(name="Site Admin", username="admin", role="admin")
            admin.set_password("admin123")
            db.session.add(admin)
            hse = User(name="HSE Officer", username="hse", role="hse")
            hse.set_password("hse123")
            db.session.add(hse)
            worker = User(name="Site Worker", username="worker", role="user")
            worker.set_password("worker123")
            db.session.add(worker)
            db.session.commit()
    app.run(debug=os.environ.get("FLASK_DEBUG", "1") == "1", host="0.0.0.0", port=5000)
